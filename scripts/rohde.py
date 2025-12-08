#==============================================================================================
# ________  ________  ________  _______   ________  _________  ___  ___  ________  ________     
#|\   ____\|\   __  \|\   __  \|\  ___ \ |\   __  \|\___   ___\\  \|\  \|\   __  \|\   __  \    
#\ \  \___|\ \  \|\  \ \  \|\ /\ \   __/|\ \  \|\  \|___ \  \_\ \  \\\  \ \  \|\  \ \  \|\  \   
# \ \  \    \ \  \\\  \ \   __  \ \  \_|/_\ \   _  _\   \ \  \ \ \  \\\  \ \   _  _\ \   __  \  
#  \ \  \____\ \  \\\  \ \  \|\  \ \  \_|\ \ \  \\  \|   \ \  \ \ \  \\\  \ \  \\  \\ \  \ \  \ 
#   \ \_______\ \_______\ \_______\ \_______\ \__\\ _\    \ \__\ \ \_______\ \__\\ _\\ \__\ \__\
#    \|_______|\|_______|\|_______|\|_______|\|__|\|__|    \|__|  \|_______|\|__|\|__|\|__|\|__|
#==============================================================================================
#==============================================================================================

#==============================================================================================
# Imports & Library Verification
#==============================================================================================
from pathlib import Path
import sys

# Verify required libraries are available
REQUIRED_LIBRARIES = {
    "pandas": "pd",
    "numpy": "np",
    "simplekml": "simplekml",
    "matplotlib": "matplotlib.pyplot",
    "openpyxl": "openpyxl"
}

def verify_libraries():
    """Check that all required libraries are installed."""
    missing = []
    for lib_name, import_alias in REQUIRED_LIBRARIES.items():
        try:
            __import__(lib_name)
        except ImportError:
            missing.append(lib_name)
    
    if missing:
        print(f"❌ ERROR: Las siguientes librerías requeridas no están instaladas:")
        for lib in missing:
            print(f"  - {lib}")
        print(f"\nInstale las librerías faltantes con:")
        print(f"  pip install {' '.join(missing)}")
        sys.exit(1)

verify_libraries()

import pandas as pd # type: ignore
import numpy as np # type: ignore
import math
import simplekml  # type: ignore
import matplotlib.pyplot as plt # type: ignore
import warnings
import time
from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font  # type: ignore

warnings.filterwarnings("ignore", category=UserWarning)

#==============================================================================================
# Rutas base
#==============================================================================================
BASE = Path(__file__).resolve().parent
DATOS = BASE.parent / "datos"      # Estructura: proyecto/ {datos, scripts, salidas}
SALIDAS = BASE.parent / "salidas"

#==============================================================================================
# Constantes de negocio
#==============================================================================================
OPERADOR_OBJETIVO = "Unión Temporal Colombia Móvil - Colombia Telecomunicaciones"
# Tabla de bandas → operador → MNC
"""CSV de referencia: tabla_mnc.csv
Banda Inicial,Banda Final,Operador,MNC
703,723,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
723.01,733,PARTNERS TELECOM COLOMBIA SAS,360
733.01,743,COMUNICACION CELULAR SA COMCEL SA,101
758,778,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
778.01,788,PARTNERS TELECOM COLOMBIA SAS,360
788.01,798,COMUNICACION CELULAR SA COMCEL SA,101
824,836.4,COMUNICACION CELULAR SA COMCEL SA,101
836.5,849,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
869,881.45,COMUNICACION CELULAR SA COMCEL SA,101
881.5,894,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
1710,1725,PARTNERS TELECOM COLOMBIA SAS,360
1725.01,1755,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
1850,1865,COMUNICACION CELULAR SA COMCEL SA,101
1867.5,1910,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
1930,1945,COMUNICACION CELULAR SA COMCEL SA,101
1947.5,1990,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
2110,2125,PARTNERS TELECOM COLOMBIA SAS,360
2125.01,2155,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
2515,2525,PARTNERS TELECOM COLOMBIA SAS,360
2525.01,2560,COMUNICACION CELULAR SA COMCEL SA,101
2635,2645,PARTNERS TELECOM COLOMBIA SAS,360
2645.01,2680,COMUNICACION CELULAR SA COMCEL SA,101
3300,3380,Unión Temporal Colombia Móvil - Colombia Telecomunicaciones,103
3380.01,3460,PARTNERS TELECOM COLOMBIA SAS,360
3460.01,3540,ESTADO,111
3540.01,3620,COMUNICACION CELULAR SA COMCEL SA,101
"""
# Definir la tabla de bandas como DataFrame
data = {
    'Banda Inicial': [703, 723.01, 733.01, 758, 778.01, 788.01, 824, 836.5, 869, 881.5, 1710, 1725.01, 1850, 1867.5, 1930, 1947.5, 2110, 2125.01, 2515, 2525.01, 2635, 2645.01, 3300, 3380.01, 3460.01, 3540.01],
    'Banda Final': [723, 733, 743, 778, 788, 798, 836.4, 849, 881.45, 894, 1725, 1755, 1865, 1910, 1945, 1990, 2125, 2155, 2525, 2560, 2645, 2680, 3380, 3460, 3540, 3620],
    'Operador': [
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'COMUNICACION CELULAR SA COMCEL SA', 
        'Unión Temporal Colombia Móvil - Colombia Telecomunicaciones', 
        'PARTNERS TELECOM COLOMBIA SAS', 
        'ESTADO',
        'COMUNICACION CELULAR SA COMCEL SA'
    ],
    'MNC': [103, 360, 101, 103, 360, 101, 101, 103, 101, 103, 360, 103, 101, 103, 101, 103, 360, 103, 360, 101, 360, 101, 103, 360, 111, 101]
}

# Crear el DataFrame
print(len(data['Banda Inicial']), len(data['Banda Final']), len(data['Operador']), len(data['MNC']))
tabla_mnc = pd.DataFrame(data)

# Tabla de bandas → operador → MNC
"""
tabla_mnc = pd.DataFrame({
    "Banda Inicial": [
        703, 723.01, 733.01, 758, 778.01, 788.01, 824, 836.51, 869, 881.51,
        1710, 1725.01, 1850, 1867.5, 1930, 1947.5, 2110, 2125.01,
        2515, 2525.01, 2635, 2645.01, 3300, 3380, 3460, 3540
    ],
    "Banda Final": [
        723, 733, 743, 778, 788, 798, 836.5, 849, 881.5, 894,
        1725, 1755, 1865, 1910, 1945, 1990, 2125, 2155,
        2525, 2555, 2645, 2675, 3379, 3450, 3530, 3610
    ],
    "Operador": [
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Partner Telecom Colombia",
        "Comcel S.A.",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Partner Telecom Colombia",
        "Comcel S.A.",
        "Comcel S.A.",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Comcel S.A.",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Partner Telecom Colombia",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Comcel S.A.",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Comcel S.A.",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Partner Telecom Colombia",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Partner Telecom Colombia",
        "Comcel S.A.",
        "Partner Telecom Colombia",
        "Comcel S.A.",
        "UT Colombia Móvil - Colombia Telecomunicaciones",
        "Comcel S.A.",
        "ESTADO",
        "UT Colombia Móvil - Colombia Telecomunicaciones"
    ],
    "MNC": [
        103, 360, 101, 103, 360, 101, 101, 103, 101, 103,
        360, 103, 101, 103, 101, 103, 360, 103,
        360, 101, 360, 101, 103, 360, 111, 101
    ]
})
"""

# Mapa MCC-MNC → PRSTM
mapa_prstm = {
    "732101": "COMUNICACION CELULAR SA COMCEL SA",
    "732103": "Unión Temporal Colombia Móvil - Colombia Telecomunicaciones",
    "732123": "Unión Temporal Colombia Móvil - Colombia Telecomunicaciones",
    "732360": "PARTNERS TELECOM COLOMBIA SAS",
}

#==============================================================================================
# Helpers de archivos
#==============================================================================================
def ensure_dirs(path: Path):
    Path(path).mkdir(parents=True, exist_ok=True)

def safe_save_generic(path: Path) -> Path:
    path = Path(path)
    if not path.exists():
        return path
    base = path.stem
    ext = path.suffix
    i = 1
    while True:
        new_path = path.parent / f"{base}_{i}{ext}"
        if not new_path.exists():
            return new_path
        i += 1

def safe_save_excel(df: pd.DataFrame, path: Path) -> Path:
    final_path = safe_save_generic(path)
    df.to_excel(final_path, index=False)
    print(f"✔ Excel guardado: {final_path}")
    return final_path

#==============================================================================================
# Lectura de CSV
#==============================================================================================
def leer_y_preparar(ruta_csv: Path) -> pd.DataFrame | None:
    try:
        df = pd.read_csv(
            ruta_csv,
            sep=",",
            encoding="utf8",
            on_bad_lines="skip",
            skiprows=2
        )
    except Exception as e:
        print(f"  ✘ Error leyendo {ruta_csv} -> {e}")
        return None

    df.columns = [c.strip() for c in df.columns]
    return df

#==============================================================================================
# Reconstrucción de GlobalCellId
#==============================================================================================
def reconstruir_globalcellid(df: pd.DataFrame, tech: str, tabla_mnc: pd.DataFrame,
                             max_ventana: int = 50) -> pd.DataFrame:
    """
    Reconstruye GlobalCellId cuando está vacío, asignando MNC correcto según frecuencia.
    Formato: '{tech} 732/{MNC}/{Frequency}/R'
    """
    df = df.copy()

    # Validar que las columnas requeridas existan
    try:
        idx_global = df.columns.get_loc("GlobalCellId")
        idx_pci = df.columns.get_loc("PCI")
        idx_freq = df.columns.get_loc("Frequency [MHz]")
    except KeyError as e:
        print(f"❌ Error: Columna requerida no encontrada: {e}")
        return df

    # Convertir frecuencia a numérico
    df.iloc[:, idx_freq] = pd.to_numeric(df.iloc[:, idx_freq], errors="coerce")

    n = len(df)
    for ii in range(n):
        pci_value = df.iat[ii, idx_pci]
        freq_value = df.iat[ii, idx_freq]
        global_val = df.iat[ii, idx_global]

        # Si GlobalCellId está vacío, intentar reconstruir
        if pd.isna(global_val) or global_val == "":
            bandas_validas = tabla_mnc[
                (tabla_mnc["Banda Inicial"] <= freq_value) &
                (freq_value <= tabla_mnc["Banda Final"])
            ]

            if not bandas_validas.empty:
                mnc_match = int(bandas_validas.iloc[0]["MNC"])
                mnc_str = f"{mnc_match:03d}"
            else:
                mnc_str = "000"

            nuevo_global = f"{tech} 732/{mnc_str}/{pci_value}/{freq_value}-R"
            df.iat[ii, idx_global] = nuevo_global

        # Copiar GlobalCellId de filas siguientes con mismo PCI
        for jj in range(ii + 1, min(ii + 1 + max_ventana, n)):
            if df.iat[jj, idx_pci] == pci_value and df.iat[jj, idx_freq] == freq_value and pd.notna(df.iat[jj, idx_global]):
                df.iat[ii, idx_global] = df.iat[jj, idx_global]
                break

    return df

#==============================================================================================
# Fecha y hora
#==============================================================================================
def obtener_fecha_hora(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y %H:%M:%S.%f")
    df["Fecha"] = df["Date"].dt.strftime("%m/%d/%Y")
    df["Hora"] = df["Date"].dt.time

    df = df.drop(columns=["Date"])
    columnas = ["Fecha", "Hora"] + [col for col in df.columns if col not in ["Fecha", "Hora"]]
    df = df[columnas]
    return df

#==============================================================================================
# Renombrar columnas
#==============================================================================================
def renombrar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={
        "Latitude": "Latitud",
        "Longitude": "Longitud",
        "Date": "Fecha/hora"
    })

#==============================================================================================
# RSRP y RSRQ (mejor valor por fila)
#==============================================================================================
def establecer_RSRP_RSRQ(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    cols_potencia = [
        "LTE NB-RSRP_Top1_byPower [dBm]",
        "LTE NB-RSRP_Top2_byPower [dBm]",
        "LTE NB-RSRP_Top3_byPower [dBm]",
    ]
    cols_calidad = [
        "LTE NB-RSRQ_Top1_byPower [dB]",
        "LTE NB-RSRQ_Top2_byPower [dB]",
        "LTE NB-RSRQ_Top3_byPower [dB]",
    ]

    df["RSRP (dBm)"] = df[cols_potencia].max(axis=1)
    df["RSRQ (dB)"] = df[cols_calidad].max(axis=1)

    df = df.drop(columns=cols_potencia + cols_calidad)
    return df

#==============================================================================================
# Operaciones sobre GlobalCellId
#==============================================================================================
def extraer_tecnologia(globalcellid: str) -> str | None:
    """Extrae la tecnología del GlobalCellId (ej: 'LTE', 'UMTS')."""
    if pd.isna(globalcellid):
        return None
    try:
        return str(globalcellid).split(" ")[0]
    except Exception:
        return None


def extraer_mcc_mnc(globalcellid: str) -> str | None:
    """Extrae el MCC-MNC del GlobalCellId."""
    if pd.isna(globalcellid):
        return None
    try:
        partes = str(globalcellid).split(" ")[1].split("/")
        return partes[0] + partes[1]
    except Exception:
        return None


def extraer_globalcell_pci(globalcellid: str) -> str | None:
    """Extrae el Global CellId y PCI del GlobalCellId."""
    if pd.isna(globalcellid):
        return None
    try:
        partes = str(globalcellid).split(" ")[1].split("/")
        return partes[2] + "/" + partes[3]
    except Exception:
        return None

#==============================================================================================
# PRSTM (operador)
#==============================================================================================
def establecer_prstm(df: pd.DataFrame) -> pd.DataFrame:
    """Asigna el operador (PRSTM) según el MCC-MNC."""
    df = df.copy()
    df["MCC-MNC"] = (
        df["MCC-MNC"].astype(str).str.replace(r"\D", "", regex=True)
    )
    df["PRSTM"] = df["MCC-MNC"].map(mapa_prstm).fillna("Desconocido")
    return df

#==============================================================================================
# Banda (MHz)
#==============================================================================================
def asignar_banda(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    f = df["Frequency [MHz]"]

    condiciones = [
        (f >= 700) & (f < 800),
        (f >= 820) & (f < 900),
        (f >= 1710) & (f < 1756),   # ← Banda 1700 MHz agregada
        (f >= 1850) & (f < 1991),
        (f >= 2100) & (f < 2156),
        (f >= 2500) & (f < 2700),
        (f >= 3300) & (f < 3700),
    ]

    bandas = [
        "700 MHz",
        "850 MHz",
        "1700 MHz",   # ← Banda agregada
        "1900 MHz",
        "2100 MHz",
        "2600 MHz",
        "3500 MHz",
    ]

    df["Banda (MHz)"] = np.select(condiciones, bandas, default="Otra banda")
    return df


#==============================================================================================
# Punto de medida (agrupación por Lat/Long)
#==============================================================================================
def asignar_punto_medida(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    cambio_punto = (
        (df["Latitud"] != df["Latitud"].shift()) |
        (df["Longitud"] != df["Longitud"].shift())
    )
    df["Punto de medida"] = cambio_punto.cumsum()
    return df

#==============================================================================================
# Eliminar columnas no necesarias
#==============================================================================================
def eliminar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["Altitude [m]", "Source"]:
        if col in df.columns:
            df = df.drop(columns=[col])
    return df

#==============================================================================================
# Marcar mejor muestra por punto (muestra conservada vs desechada)
#==============================================================================================
def marcar_mejor_muestra(df: pd.DataFrame) -> pd.DataFrame:
    """
    Marca con 'muestra conservada' la mejor muestra por punto de medición
    (mayor RSRP), y marca las demás como desechadas.
    El DF ya debe venir filtrado por el operador objetivo.
    """
    df = df.copy()

    df["Exclusiones"] = "muestra desechada por haber un mayor RSRP para UT en este punto"

    # Vectorized groupby with idxmax - much faster than iterrows
    idx_mejores = df.groupby("Punto de medida")["RSRP (dBm)"].idxmax()

    # Vectorized assignment
    df.loc[idx_mejores, "Exclusiones"] = " "

    return df

#==============================================================================================
# Clasificación por rangos de RSRP (centralizado)
#==============================================================================================
def clasificar_rsrp(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega la columna 'color_rsrp' con valores:
    - red, yellow, green, blue o None
    según el rango de RSRP.
    """
    df = df.copy()
    rsrp = df["RSRP (dBm)"]

    condiciones = [
        (rsrp >= -80) & (rsrp <= -30),
        (rsrp >= -90) & (rsrp < -80),
        (rsrp >= -100) & (rsrp < -90),
        (rsrp >= -150) & (rsrp < -100),
    ]
    colores = ["red", "yellow", "green", "blue"]

    df["color_rsrp"] = np.select(condiciones, colores, default=None)
    return df
    
#==============================================================================================
# Clasificación por rangos de RSRQ
#==============================================================================================
def clasificar_rsrq(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clasificación de calidad
    """
    df = df.copy()
    rsrq = df["RSRQ (dB)"]
    
    condiciones = [
        (rsrq >= -10) & (rsrq <= 0),
        (rsrq >= -15) & (rsrq < -10.001),
        (rsrq >= -20) & (rsrq < -15.001),
        (rsrq >= -40) & (rsrq < -20.001)
    ]
    colores = ["red", "yellow", "green", "blue"]
    df["color_rsrq"] = np.select(condiciones, colores, default=None)
    return df
    
#==============================================================================================
# Resumen RSRP
#==============================================================================================
def generar_resumen_rsrp(df: pd.DataFrame) -> pd.DataFrame:
    df_4g = df[df["Tecnología"] == "LTE"].copy()

    # Vectorized operations for counting - much faster than len()
    rsrp_col = df_4g["RSRP (dBm)"]
    total_muestras_4g = len(df_4g)
    buenas_4g = (rsrp_col >= -100).sum()   # Vectorized
    malas_4g = (rsrp_col < -100).sum()     # Vectorized

    #df_validas = df_4g[df_4g["Exclusiones"] == "muestra conservada"].copy()
    valid_mask = (df_4g["Exclusiones"] == " ") & (df_4g["dentro_2km"] == " ")
    df_validas = df_4g[valid_mask].copy()
    total_validas = len(df_validas)
    buenas_validas = (df_validas["RSRP (dBm)"] >= -100).sum()  # Vectorized
    malas_validas = len(df_validas[df_validas["RSRP (dBm)"] < -100])
    
    """
    # Obtener bandas de frecuencia para el resumen
    for banda in df_4g["Banda (MHz)"].unique():
        # Poner los nombres de las bandas en el resumen concadenados
        bandas_str = ", ".join(sorted(df_4g["Banda (MHz)"].unique().astype(str)))
    """
    bandas_str = ", ".join(sorted(df_4g["Banda (MHz)"].unique().astype(str)))
    muestras_excluidas = total_muestras_4g - total_validas
    resumen = pd.DataFrame({
        f"Niveles de señal RSRP Banda {bandas_str}": [        #modificar
            "Muestras 4G",
            "RSRP ≥ -100 dBm",
            "RSRP < -100 dBm",
            "Muestras excluidas"
        ],
        "Total Muestras": [
            total_muestras_4g,
            buenas_4g,
            malas_4g,
            muestras_excluidas
        ],
        "Porcentaje (%)": [
            "100.00%" if total_muestras_4g > 0 else "0.00%",
            f"{(buenas_4g / total_muestras_4g) * 100:.2f}%" if total_muestras_4g > 0 else "0.00%",
            f"{(malas_4g / total_muestras_4g) * 100:.2f}%" if total_muestras_4g > 0 else "0.00%",
            " "
        ],
        "Muestras válidas Radio 2Km desde PMCP": [
            total_validas,
            buenas_validas,
            malas_validas,
            " "
        ],
        "% muestras válidas 4G": [
            "100.00%" if total_validas > 0 else "0.00%",
            f"{(buenas_validas / total_validas) * 100:.2f}%" if total_validas > 0 else "0.00%",
            f"{(malas_validas / total_validas) * 100:.2f}%" if total_validas > 0 else "0.00%",
            " "
        ],
    })

    return resumen

def exportar_excel_con_resumen(df: pd.DataFrame, resumen: pd.DataFrame, ruta_salida: Path):
    """
    Exporta datos y resumen a Excel con estilos profesionales:
    - Bordes en todas las celdas
    - Cabeceras con color azul claro (94, 171, 219)
    - Filas de datos con fondo azul claro (#BDD7EE)
    - Columnas de df_procesar con encabezados amarillos
    - Ancho de columnas automático
    """
    ruta_final = safe_save_generic(ruta_salida)
    
    # Escribir con xlsxwriter primero
    with pd.ExcelWriter(ruta_final, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Datos Procesados", index=False)
        resumen.to_excel(writer, sheet_name="Resumen RSRP", index=False)
    
    # Ahora abrir con openpyxl para aplicar estilos
    try:
        from openpyxl.styles import Border, Side, PatternFill, Alignment, Font  # type: ignore
        from openpyxl import load_workbook  # type: ignore
        
        wb = load_workbook(ruta_final)
        
        # Estilos comunes
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Color azul claro para cabeceras: RGB(94, 171, 219)
        header_fill = PatternFill(start_color="5EABDB", end_color="5EABDB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Color azul claro para celdas de datos: #BDD7EE
        data_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # ════════════════════════════════════════════════════════════════════════════
        # Aplicar estilos a hoja "Resumen RSRP"
        # ════════════════════════════════════════════════════════════════════════════
        ws_resumen = wb["Resumen RSRP"]
        
        # Aplicar bordes y formato a todas las celdas con datos
        for row in ws_resumen.iter_rows(min_row=1, max_row=ws_resumen.max_row,
                                        min_col=1, max_col=ws_resumen.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                
                # Cabecera (primera fila)
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    # Aplicar color de fondo a todas las filas de datos
                    cell.fill = data_fill
        
        # Ajustar ancho de columnas automáticamente
        for col in ws_resumen.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)  # Máximo 50 caracteres
            ws_resumen.column_dimensions[col_letter].width = adjusted_width
        
        # Establecer altura de fila para cabecera
        ws_resumen.row_dimensions[1].height = 30
        
        # ════════════════════════════════════════════════════════════════════════════
        # Aplicar estilos a hoja "Datos Procesados" - encabezados amarillos
        # ════════════════════════════════════════════════════════════════════════════
        ws_datos = wb["Datos Procesados"]
        
        # Color amarillo para encabezados de datos
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        yellow_font = Font(bold=True, color="000000")
        
        # Aplicar estilos a la fila de encabezados (primera fila)
        for cell in ws_datos[1]:
            cell.border = thin_border
            cell.fill = yellow_fill
            cell.font = yellow_font
            cell.alignment = center_align
        
        # Aplicar bordes a todas las otras celdas
        for row in ws_datos.iter_rows(min_row=2, max_row=ws_datos.max_row,
                                       min_col=1, max_col=ws_datos.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Ajustar ancho de columnas automáticamente
        for col in ws_datos.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)  # Máximo 40 caracteres
            ws_datos.column_dimensions[col_letter].width = adjusted_width
        
        # Establecer altura de fila para cabecera
        ws_datos.row_dimensions[1].height = 30
        
        wb.save(ruta_final)
        print(f"✔ Excel con resumen guardado: {ruta_final}")
        
    except ImportError:
        print(f"⚠ Advertencia: openpyxl no disponible. Excel guardado sin estilos avanzados.")
        print(f"  Instale openpyxl con: pip install openpyxl")
    
def generar_kmz_completo(df: pd.DataFrame, CODIGO: str, LOCALIDAD: str, MNC_SELECCIONADOS: list):
    """
    Genera un ÚNICO archivo KMZ que contiene:
    - Radio 2 km
    - Potencia (RSRP)
    - Calidad (RSRQ)
    - Bandas
    - MNC
    - GlobalCell/PCI
    """

    SALIDAS.mkdir(parents=True, exist_ok=True)

    #=====================================================================
    # Crear un solo objeto KML
    #=====================================================================
    kml = simplekml.Kml()

    ICON_URL = "http://maps.google.com/mapfiles/kml/shapes/shaded_dot.png"

    #=====================================================================
    # 1) RADIO 2 KM
    #=====================================================================
    folder_radio = kml.newfolder(name="Círculo 2 km + puntos origen")

    # Visibilidad con google earth
    #folder_radio.style.liststyle.listitemtype = "check"   # visible por defecto

    archivo_coord = DATOS / "coordenadas_origen.txt"
    if archivo_coord.exists():
        try:
            puntos = []
            with open(archivo_coord, "r", encoding="utf-8") as f:
                for linea in f:
                    linea = linea.strip()
                    if not linea:
                        continue
                    partes = [x.strip() for x in linea.split(",")]
                    if len(partes) != 3:
                        continue
                    nombre, lat, lon = partes
                    puntos.append((nombre, float(lat), float(lon)))

            if puntos:
                nombre_origen, lat0, lon0 = puntos[0]

                R = 6371000
                radio_m = 2000
                coords_circulo = []

                for ang in range(0, 361):
                    brng = math.radians(ang)
                    lat_rad = math.radians(lat0)
                    lon_rad = math.radians(lon0)
                    lat_c = math.asin(
                        math.sin(lat_rad) * math.cos(radio_m / R) +
                        math.cos(lat_rad) * math.sin(radio_m / R) * math.cos(brng)
                    )
                    lon_c = lon_rad + math.atan2(
                        math.sin(brng) * math.sin(radio_m / R) * math.cos(lat_rad),
                        math.cos(radio_m / R) - math.sin(lat_rad) * math.sin(lat_c)
                    )
                    coords_circulo.append((math.degrees(lon_c), math.degrees(lat_c)))

                coords_circulo.append(coords_circulo[0])

                pol = folder_radio.newpolygon(
                    name=f"Círculo 2 km alrededor de {nombre_origen}",
                    outerboundaryis=coords_circulo,
                )
                pol.style.polystyle.fill = 0
                pol.style.linestyle.color = simplekml.Color.hex("ff00ff")
                pol.style.linestyle.width = 3

                ICON22 = "http://maps.google.com/mapfiles/kml/shapes/donut.png"
                for nombre, lat, lon in puntos:
                    p = folder_radio.newpoint(name=nombre, coords=[(lon, lat)])
                    p.style.labelstyle.color = simplekml.Color.orange
                    p.style.iconstyle.color = simplekml.Color.hex("ff00ff")
                    p.style.iconstyle.icon.href = ICON22
                    p.style.iconstyle.scale = 1.2

        except Exception as e:
            print(f"⚠ Error procesando coordenadas: {e}")

    #=====================================================================
    # 2) POTENCIA (RSRP)
    #=====================================================================
    df_rsrp = clasificar_rsrp(df.copy())
    df_rsrp = df_rsrp[df_rsrp["color_rsrp"].notna()]
    folder_rsrp = kml.newfolder(name="Potencia (RSRP)")

    folders_potencia = {
        "red":    folder_rsrp.newfolder(name="-80 a -30 dBm"),
        "yellow": folder_rsrp.newfolder(name="-90 a -80 dBm"),
        "green":  folder_rsrp.newfolder(name="-100 a -90 dBm"),
        "blue":   folder_rsrp.newfolder(name="-150 a -100 dBm"),
    }

    for _, row in df_rsrp.iterrows():
        color = row["color_rsrp"]
        f = folders_potencia[color]
        p = f.newpoint(coords=[(row["Longitud"], row["Latitud"])])
        #p.name = f"{row['RSRP (dBm)']:.0f} dBm"
        p.description = "\n".join(f"{c}: {v}" for c, v in row.items())
        p.style.iconstyle.color = getattr(simplekml.Color, color)
        p.style.iconstyle.icon.href = ICON_URL
        p.style.iconstyle.scale = 0.6

    #=====================================================================
    # 3) CALIDAD (RSRQ)
    #=====================================================================
    df_rsrq = clasificar_rsrq(df.copy())
    df_rsrq = df_rsrq[df_rsrq["color_rsrq"].notna()]
    folder_rsrq = kml.newfolder(name="Calidad (RSRQ)")

    folders_calidad = {
        "red":    folder_rsrq.newfolder(name="-10 a 0 dB"),
        "yellow": folder_rsrq.newfolder(name="-15 a -10 dB"),
        "green":  folder_rsrq.newfolder(name="-20 a -15 dB"),
        "blue":   folder_rsrq.newfolder(name="-40 a -20 dB"),
    }
    
    for _, row in df_rsrq.iterrows():
        color = row["color_rsrq"]
        f = folders_calidad[color]
        p = f.newpoint(coords=[(row["Longitud"], row["Latitud"])])
        #p.name = f"{row['RSRQ (dB)']:.0f} dB"
        p.description = "\n".join(f"{c}: {v}" for c, v in row.items())
        p.style.iconstyle.color = getattr(simplekml.Color, color)
        p.style.iconstyle.icon.href = ICON_URL
        p.style.iconstyle.scale = 0.6

    #=====================================================================
    # 4) BANDAS
    #=====================================================================
    df_banda = df[df["Banda (MHz)"].notna()]
    folder_banda = kml.newfolder(name="Bandas")   

    COLOR_BANDA_KML = {
        "700 MHz":  simplekml.Color.rgb(255, 0, 0, 255),
        "850 MHz":  simplekml.Color.rgb(255, 165, 0, 255),
        "1700 MHz": simplekml.Color.rgb(0, 255, 255, 255),
        "1900 MHz": simplekml.Color.rgb(255, 255, 0, 255),
        "2100 MHz": simplekml.Color.rgb(0, 128, 0, 255),
        "2600 MHz": simplekml.Color.rgb(0, 0, 255, 255),
        "3500 MHz": simplekml.Color.rgb(128, 0, 128, 255),
        "Otra banda": simplekml.Color.rgb(128, 128, 128, 255),
    }

    carpetas_banda = {
        b: folder_banda.newfolder(name=f"Banda {b}") for b in df_banda["Banda (MHz)"].unique()
    }

    for _, row in df_banda.iterrows():
        b = row["Banda (MHz)"]
        p = carpetas_banda[b].newpoint(coords=[(row["Longitud"], row["Latitud"])])
        #p.name = str(b)
        p.description = "\n".join(f"{c}: {v}" for c, v in row.items())
        p.style.iconstyle.icon.href = ICON_URL
        p.style.iconstyle.color = COLOR_BANDA_KML.get(b, simplekml.Color.gray)
        p.style.iconstyle.scale = 0.6

    #=====================================================================
    # 5) MNC
    #=====================================================================
    df_mnc = df[df["MCC-MNC"].notna()]
    folder_mnc = kml.newfolder(name="MNC")  
    
    colores_disponibles = [
        simplekml.Color.blue,
        simplekml.Color.lime,
        simplekml.Color.red,
        simplekml.Color.orange,
    ]

    COLOR_MNC = {}
    for i, m in enumerate(MNC_SELECCIONADOS):
        COLOR_MNC[f"732{m}"] = colores_disponibles[i % len(colores_disponibles)]

    carpetas_mnc = {
        m: folder_mnc.newfolder(name=f"MNC {m}") for m in df_mnc["MCC-MNC"].unique()
    }

    for _, row in df_mnc.iterrows():
        m = row["MCC-MNC"]
        p = carpetas_mnc[m].newpoint(coords=[(row["Longitud"], row["Latitud"])])
        #p.name = str(m)
        p.description = "\n".join(f"{c}: {v}" for c, v in row.items())
        p.style.iconstyle.icon.href = ICON_URL
        p.style.iconstyle.color = COLOR_MNC.get(m, simplekml.Color.gray)
        p.style.iconstyle.scale = 0.6

    #=====================================================================
    # 6) GLOBALCELL / PCI
    #=====================================================================
    df_gc = df[df["Global CellId /PCI"].notna()].copy()
    folder_gc = kml.newfolder(name="GlobalCell / PCI")
 
    vc = df_gc["Global CellId /PCI"].value_counts()
    top_vals = list(vc.head(5).index)

    def cat(v): return v if v in top_vals else "Otros"
    df_gc["CAT"] = df_gc["Global CellId /PCI"].apply(cat)

    colores_gc = [
        simplekml.Color.red,
        simplekml.Color.blue,
        simplekml.Color.lime,
        simplekml.Color.yellow,
        simplekml.Color.purple,
    ]

    COLOR_GC = {val: colores_gc[i] for i, val in enumerate(top_vals)}
    COLOR_GC["Otros"] = simplekml.Color.gray

    carpetas_gc = {
        c: folder_gc.newfolder(name=c) for c in df_gc["CAT"].unique()
    }

    for _, row in df_gc.iterrows():
        c = row["CAT"]
        p = carpetas_gc[c].newpoint(coords=[(row["Longitud"], row["Latitud"])])
        #p.name = str(row["Global CellId /PCI"])
        p.description = "\n".join(f"{col}: {val}" for col, val in row.items())
        p.style.iconstyle.icon.href = ICON_URL
        p.style.iconstyle.color = COLOR_GC[c]
        p.style.iconstyle.scale = 0.6

    #=====================================================================
    # GUARDAR KMZ ÚNICO
    #=====================================================================
    ruta_kmz = SALIDAS / f"{CODIGO} {LOCALIDAD}" / f"{CODIGO} {LOCALIDAD}.kmz"
    ruta_kmz = safe_save_generic(ruta_kmz)
    kml.savekmz(str(ruta_kmz))

    print(f"✔ KMZ ÚNICO generado: {ruta_kmz}")


#==============================================================================================
# Mapa RSRP con matplotlib
#==============================================================================================
"""
def generar_mapa_rsrp(df_conservada: pd.DataFrame, CODIGO: str, LOCALIDAD: str):
    SALIDAS.mkdir(parents=True, exist_ok=True)

    df_conservada = clasificar_rsrp(df_conservada)
    df_plot = df_conservada[df_conservada["color_rsrp"].notna()].copy()

    count_red = (df_plot["color_rsrp"] == "red").sum()
    count_yellow = (df_plot["color_rsrp"] == "yellow").sum()
    count_green = (df_plot["color_rsrp"] == "green").sum()
    count_blue = (df_plot["color_rsrp"] == "blue").sum()

    plt.figure(figsize=(10, 10))

    for color, label, count in [
        ("red", "Rojo (-80 a -30 dBm)", count_red),
        ("yellow", "Amarillo (-90 a -80 dBm)", count_yellow),
        ("green", "Verde (-100 a -90 dBm)", count_green),
        ("blue", "Azul (-150 a -100 dBm)", count_blue),
    ]:
        df_c = df_plot[df_plot["color_rsrp"] == color]
        if not df_c.empty:
            plt.scatter(
                df_c["Longitud"],
                df_c["Latitud"],
                s=8,
                c=color,
                label=f"{label} → {count} muestras",
            )

    total = len(df_plot)
    plt.title(f"Mapa RSRP - {CODIGO} {LOCALIDAD}\nTotal muestras: {total}", fontsize=14)

    plt.axis("off")
    plt.legend(loc="upper right")
    plt.margins(0)
    plt.tight_layout()

    ruta_png = SALIDAS / f"{CODIGO} {LOCALIDAD}"/ f"{CODIGO} {LOCALIDAD} RSRP.png"
    ruta_png = safe_save_generic(ruta_png)
    plt.savefig(str(ruta_png), dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close()
 """  

def generar_mapa_rsrp(
    df_conservada: pd.DataFrame, 
    CODIGO: str, 
    LOCALIDAD: str, 
    lat0: float, lon0: float, nombre0: str,   # primer punto
    lat1: float, lon1: float, nombre1: str,  # segundo punto
    radio_m: float = 2000       # radio en metros
):
    SALIDAS.mkdir(parents=True, exist_ok=True)

    df_conservada = clasificar_rsrp(df_conservada)
    df_plot = df_conservada[df_conservada["color_rsrp"].notna()].copy()

    # Contar colores
    count_red = (df_plot["color_rsrp"] == "red").sum()
    count_yellow = (df_plot["color_rsrp"] == "yellow").sum()
    count_green = (df_plot["color_rsrp"] == "green").sum()
    count_blue = (df_plot["color_rsrp"] == "blue").sum()

    plt.figure(figsize=(10, 10))

    # Dibujar puntos por color
    handles = []
    for color, label, count in [
        ("red", "(-30 a -80 dBm)", count_red),
        ("yellow", "(-80 a -90 dBm)", count_yellow),
        ("green", "(-90 a -100 dBm)", count_green),
        ("blue", "(-100 a -150 dBm)", count_blue),
    ]:
        df_c = df_plot[df_plot["color_rsrp"] == color]
        if not df_c.empty:
            scatter = plt.scatter(
                df_c["Longitud"],
                df_c["Latitud"],
                s=50,
                c=color,
                label=f"{label} → {count} muestras"
            )
            handles.append(scatter)

    # Dibujar círculo alrededor de (lat0, lon0)
    # Aproximación: 1 grado lat ≈ 111 km, 1 grado lon ≈ 111 km * cos(lat)
    lat_radius = radio_m / 111000
    # Evitar división por cero cerca de los polos
    cos_lat = np.cos(np.radians(lat0))
    if abs(cos_lat) < 0.001:
        cos_lat = 0.001
    lon_radius = radio_m / (111000 * cos_lat)

    theta = np.linspace(0, 2 * np.pi, 360)
    circle_lat = lat0 + lat_radius * np.sin(theta)
    circle_lon = lon0 + lon_radius * np.cos(theta)
    plt.plot(circle_lon, circle_lat, color="magenta", linewidth=2)

    # Dibujar puntos específicos
    plt.scatter([lon0], [lat0], color="black", marker="^", s=80)
    plt.scatter([lon1], [lat1], color="black", marker="^", s=80)
    
    # Colocar el nombre en los puntos
    plt.annotate(nombre0, (lon0, lat0), textcoords="offset points",
                 xytext=(-15, -15), ha='right', va='top',
                 fontsize=10, fontweight='bold', color='white',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8))
    plt.annotate(nombre1, (lon1, lat1), textcoords="offset points",
                 xytext=(15, 15), ha='left', va='bottom',
                 fontsize=10, fontweight='bold', color='white',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8))

    total = len(df_plot)
    plt.title(f"Mapa RSRP - {CODIGO} {LOCALIDAD}\nTotal muestras: {total}", fontsize=14)

    # Modificar leyenda
    plt.axis("off")
    plt.legend(handles=handles, loc="upper right", fontsize=14, markerscale=2)  # Cambiar markerscale para aumentar tamaño de los puntos
    plt.margins(0)
    plt.tight_layout()

    ruta_png = SALIDAS / f"{CODIGO} {LOCALIDAD}" / f"{CODIGO} {LOCALIDAD} RSRP.png"
    ruta_png = safe_save_generic(ruta_png)
    plt.savefig(str(ruta_png), dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close()

    
#==============================================================================================
# Mapa RSRQ con matplotlib
#==============================================================================================
def generar_mapa_rsrq(
    df_conservada: pd.DataFrame, 
    CODIGO: str, 
    LOCALIDAD: str, 
    lat0: float, lon0: float, nombre0: str,   # primer punto
    lat1: float, lon1: float, nombre1: str,  # segundo punto
    radio_m: float = 2000       # radio en metros
):
    SALIDAS.mkdir(parents=True, exist_ok=True)

    df_conservada = clasificar_rsrq(df_conservada)
    df_plot = df_conservada[df_conservada["color_rsrq"].notna()].copy()

    # Contar colores
    count_red = (df_plot["color_rsrq"] == "red").sum()
    count_yellow = (df_plot["color_rsrq"] == "yellow").sum()
    count_green = (df_plot["color_rsrq"] == "green").sum()
    count_blue = (df_plot["color_rsrq"] == "blue").sum()

    plt.figure(figsize=(10, 10))

    # Dibujar puntos por color
    handles = []  # Para personalizar la leyenda
    for color, label, count in [
        ("red", "(0 a -10 dB)", count_red),
        ("yellow", "(-10.001 a -15 dB)", count_yellow),
        ("green", "(-15.001 a -20 dB)", count_green),
        ("blue", "(-20.001 a -40 dB)", count_blue),
    ]:
        df_c = df_plot[df_plot["color_rsrq"] == color]
        if not df_c.empty:
            scatter = plt.scatter(
                df_c["Longitud"],
                df_c["Latitud"],
                s=50,
                c=color,
                label=f"{label} → {count} muestras"
            )
            handles.append(scatter)

    # Dibujar círculo alrededor de (lat0, lon0)
    lat_radius = radio_m / 111000
    # Evitar división por cero cerca de los polos
    cos_lat = np.cos(np.radians(lat0))
    if abs(cos_lat) < 0.001:
        cos_lat = 0.001
    lon_radius = radio_m / (111000 * cos_lat)
    theta = np.linspace(0, 2 * np.pi, 360)
    circle_lat = lat0 + lat_radius * np.sin(theta)
    circle_lon = lon0 + lon_radius * np.cos(theta)
    plt.plot(circle_lon, circle_lat, color="magenta", linewidth=2)

    # Dibujar puntos específicos
    plt.scatter([lon0], [lat0], color="black", marker="^", s=80)
    plt.scatter([lon1], [lat1], color="black", marker="^", s=80)
    
    # Colocar el nombre en los puntos
    plt.annotate(nombre0, (lon0, lat0), textcoords="offset points",
                 xytext=(-15, -15), ha='right', va='top',
                 fontsize=10, fontweight='bold', color='white',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8))
    plt.annotate(nombre1, (lon1, lat1), textcoords="offset points",
                 xytext=(15, 15), ha='left', va='bottom',
                 fontsize=10, fontweight='bold', color='white',
                 bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8))

    total = len(df_plot)
    plt.title(f"Mapa RSRQ - {CODIGO} {LOCALIDAD}\nTotal muestras: {total}", fontsize=14)

    # Modificar leyenda para aumentar tamaño de los puntos
    plt.axis("off")
    plt.legend(handles=handles, loc="upper right", fontsize=14, markerscale=2)  # Cambiar markerscale para aumentar el tamaño de los puntos
    plt.margins(0)
    plt.tight_layout()

    ruta_png = SALIDAS / f"{CODIGO} {LOCALIDAD}" / f"{CODIGO} {LOCALIDAD} RSRQ.png"
    ruta_png = safe_save_generic(ruta_png)
    plt.savefig(str(ruta_png), dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close()
    
#==============================================================================================
# Mapa Banda de frecuencia
#==============================================================================================    
def generar_mapa_banda(
    df: pd.DataFrame,
    CODIGO: str,
    LOCALIDAD: str,
    lat0: float, lon0: float, nombre0: str,
    lat1: float, lon1: float, nombre1: str,
    radio_m: float = 2000
):
    SALIDAS.mkdir(parents=True, exist_ok=True)

    df_banda = df.copy()
    df_banda = df_banda[df_banda["Banda (MHz)"].notna()]

    if df_banda.empty:
        print("⚠ No hay datos con 'Banda (MHz)' para generar el mapa de bandas.")
        return

    COLOR_BANDA_PLOT = {
        "700 MHz": "#FF0000",  
        "850 MHz": "#FF9900",
        "1700 MHz": "#EC6AB8",
        "1900 MHz": "#FFFF00",
        "2100 MHz": "#008000",
        "2600 MHz": "#0000FF",
        "3500 MHz": "#800080",
        "Otra banda": "#808080",
    }

    conteo = df_banda["Banda (MHz)"].value_counts().to_dict()

    plt.figure(figsize=(10, 10))

    handles = []
    for banda in sorted(df_banda["Banda (MHz)"].unique()):
        df_b = df_banda[df_banda["Banda (MHz)"] == banda]
        color = COLOR_BANDA_PLOT.get(banda, "#808080")
        n = conteo.get(banda, 0)
        scatter = plt.scatter(
            df_b["Longitud"],
            df_b["Latitud"],
            s=50,
            c=color,
            label=f"{banda} → {n} muestras"
        )
        handles.append(scatter)

    # Dibujar círculo
    lat_radius = radio_m / 111000
    lon_radius = radio_m / (111000 * np.cos(np.radians(lat0)))
    theta = np.linspace(0, 2 * np.pi, 360)
    circle_lat = lat0 + lat_radius * np.sin(theta)
    circle_lon = lon0 + lon_radius * np.cos(theta)
    plt.plot(circle_lon, circle_lat, color="magenta", linewidth=2)

    # Dibujar puntos específicos
    plt.scatter([lon0], [lat0], color="black", marker="^", s=80)
    plt.scatter([lon1], [lat1], color="black", marker="^", s=80)
    
    # colocar el nombre en los puntos
    plt.annotate(nombre0, (lon0, lat0), textcoords = "offset points",
        xytext = (-15, -15),
        ha = 'right',
        va = 'top',
        fontsize=10, fontweight='bold', color='white',
        bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8)
    )
    plt.annotate(nombre1, (lon1, lat1), textcoords = "offset points",
        xytext = (15, 15),
        ha = 'left',
        va = 'bottom',
        fontsize=10, fontweight='bold', color='white',
        bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8)
    )
    
    total = len(df_banda)
    plt.title(f"Mapa por Banda (MHz) - {CODIGO} {LOCALIDAD}\nTotal muestras: {total}", fontsize=14)

    plt.axis("off")
    plt.legend(handles=handles, loc="upper right", fontsize=14, markerscale=2)  # Ajuste del tamaño de los puntos en la leyenda
    plt.margins(0)
    plt.tight_layout()

    ruta_png = SALIDAS / f"{CODIGO} {LOCALIDAD}" / f"{CODIGO} {LOCALIDAD} Bandas.png"
    ruta_png = safe_save_generic(ruta_png)
    plt.savefig(str(ruta_png), dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close()

    
#==============================================================================================
# Mapa MCC-MNC con matplotlib
#==============================================================================================
def generar_mapa_mnc(
    df: pd.DataFrame, 
    CODIGO: str, 
    LOCALIDAD: str, 
    MNC_SELECCIONADOS: list,
    lat0: float, lon0: float, nombre0: str,
    lat1: float, lon1: float, nombre1: str,
    radio_m: float = 2000
):
    SALIDAS.mkdir(parents=True, exist_ok=True)
    df_mnc = df.copy()
    df_mnc = df_mnc[df_mnc["MCC-MNC"].notna()]

    if df_mnc.empty:
        print("⚠ No hay datos con 'MCC-MNC' para generar el mapa de mnc.")
        return

    # Colores para diferentes MNCs
    #colores_disponibles = ["#008000", "#0000FF", "#FF0000", "#FFA500"]
    colores_disponibles = ["#0000FF", "#008000", "#FF0000", "#FFA500"]
    COLOR_MNC_PLOT = {}
    for i, mnc in enumerate(MNC_SELECCIONADOS):
        color = colores_disponibles[i % len(colores_disponibles)]
        COLOR_MNC_PLOT[f"732{mnc}"] = color

    conteo = df_mnc["MCC-MNC"].value_counts().to_dict()

    plt.figure(figsize=(10, 10))

    handles = []
    for mnc in sorted(df_mnc["MCC-MNC"].unique()):
        df_b = df_mnc[df_mnc["MCC-MNC"] == mnc]
        color = COLOR_MNC_PLOT.get(mnc, "#808080")
        n = conteo.get(mnc, 0)
        scatter = plt.scatter(
            df_b["Longitud"],
            df_b["Latitud"],
            s=50,
            c=color,
            label=f"{mnc} → {n} muestras"
        )
        handles.append(scatter)

    # Dibujar círculo
    lat_radius = radio_m / 111000
    lon_radius = radio_m / (111000 * np.cos(np.radians(lat0)))
    theta = np.linspace(0, 2 * np.pi, 360)
    circle_lat = lat0 + lat_radius * np.sin(theta)
    circle_lon = lon0 + lon_radius * np.cos(theta)
    plt.plot(circle_lon, circle_lat, color="magenta", linewidth=2)

    # Dibujar puntos específicos
    plt.scatter([lon0], [lat0], color="black", marker="^", s=80)
    plt.scatter([lon1], [lat1], color="black", marker="^", s=80)
    
    # colocar el nombre en los puntos
    plt.annotate(nombre0, (lon0, lat0), textcoords = "offset points",
        xytext = (-15, -15),
        ha = 'right',
        va = 'top',
        fontsize=10, fontweight='bold', color='white',
        bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8)
    )
    plt.annotate(nombre1, (lon1, lat1), textcoords = "offset points",
        xytext = (15, 15),
        ha = 'left',
        va = 'bottom',
        fontsize=10, fontweight='bold', color='white',
        bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8)
    )
    
    total = len(df_mnc)
    plt.title(f"Mapa por MCC-MNC - {CODIGO} {LOCALIDAD}\nTotal muestras: {total}", fontsize=14)

    plt.axis("off")
    plt.legend(handles=handles, loc="upper right", fontsize=14, markerscale=2)  # Ajuste del tamaño de los puntos en la leyenda
    plt.margins(0)
    plt.tight_layout()

    ruta_png = SALIDAS / f"{CODIGO} {LOCALIDAD}" / f"{CODIGO} {LOCALIDAD} MNC.png"
    ruta_png = safe_save_generic(ruta_png)
    plt.savefig(str(ruta_png), dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close()

#==============================================================================================
# Mapa GLOBALCELL/PCI
#==============================================================================================     
def generar_mapa_globalcellpci(
    df: pd.DataFrame, 
    CODIGO: str, 
    LOCALIDAD: str,
    lat0: float, lon0: float, nombre0: str,
    lat1: float, lon1: float, nombre1: str,
    radio_m: float = 2000
):
    SALIDAS.mkdir(parents=True, exist_ok=True)
    df_gc = df.copy()
    
    if "Global CellId /PCI" not in df_gc.columns:
        print("⚠ No existe la columna 'Global CellId /PCI' en el DataFrame.")
        return

    df_gc = df_gc[df_gc["Global CellId /PCI"].notna()]
    if df_gc.empty:
        print("⚠ No hay datos con 'Global CellId /PCI' para generar el mapa.")
        return

    # Top 5 valores
    vc = df_gc["Global CellId /PCI"].value_counts()
    top_n = min(5, len(vc))
    top_vals = list(vc.head(top_n).index)

    def categorizar(val):
        return val if val in top_vals else "Otros"

    df_gc["GCID_cat"] = df_gc["Global CellId /PCI"].apply(categorizar)
    conteo_cat = df_gc["GCID_cat"].value_counts().to_dict()

    colores_plot = ["#FF0000", "#0000FF", "#00AA00", "#FFA500", "#800080"]
    COLOR_CAT = {val: colores_plot[i] for i, val in enumerate(top_vals)}
    COLOR_CAT["Otros"] = "#808080"

    plt.figure(figsize=(10, 10))

    handles = []
    for cat in conteo_cat.keys():
        df_c = df_gc[df_gc["GCID_cat"] == cat]
        color = COLOR_CAT.get(cat, "#808080")
        n = conteo_cat[cat]
        label = f"{cat} → {n} muestras" if cat != "Otros" else f"Otros → {n} muestras"

        scatter = plt.scatter(
            df_c["Longitud"],
            df_c["Latitud"],
            s=50,
            c=color,
            label=label
        )
        handles.append(scatter)

    # Dibujar círculo
    lat_radius = radio_m / 111000
    # Evitar división por cero cerca de los polos
    cos_lat = np.cos(np.radians(lat0))
    if abs(cos_lat) < 0.001:
        cos_lat = 0.001
    lon_radius = radio_m / (111000 * cos_lat)
    theta = np.linspace(0, 2 * np.pi, 360)
    circle_lat = lat0 + lat_radius * np.sin(theta)
    circle_lon = lon0 + lon_radius * np.cos(theta)
    plt.plot(circle_lon, circle_lat, color="magenta", linewidth=2)

    # Dibujar puntos específicos
    plt.scatter([lon0], [lat0], color="black", marker="^", s=80)
    plt.scatter([lon1], [lat1], color="black", marker="^", s=80)
    
    # colocar el nombre en los puntos
    plt.annotate(nombre0, (lon0, lat0), textcoords = "offset points",
        xytext = (-15, -15),
        ha = 'right',
        va = 'top',
        fontsize=10, fontweight='bold', color='white',
        bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8)
    )
    plt.annotate(nombre1, (lon1, lat1), textcoords = "offset points",
        xytext = (15, 15),
        ha = 'left',
        va = 'bottom',
        fontsize=10, fontweight='bold', color='white',
        bbox=dict(boxstyle='round,pad=0.2', facecolor='black', edgecolor='yellow', linewidth=1.5, alpha=0.8)
    )
    
    total = len(df_gc)
    plt.title(
        f"CID sector (top 5 + Otros) - {CODIGO} {LOCALIDAD}\nTotal muestras: {total}",
        fontsize=13
    )

    plt.axis("off")
    plt.legend(handles=handles, loc="upper right", fontsize=14, markerscale=2)  # Ajuste del tamaño de los puntos en la leyenda
    plt.margins(0)
    plt.tight_layout()

    ruta_png = SALIDAS / f"{CODIGO} {LOCALIDAD}" / f"{CODIGO} {LOCALIDAD} CID-sector.png"
    ruta_png = safe_save_generic(ruta_png)
    plt.savefig(str(ruta_png), dpi=300, bbox_inches="tight", pad_inches=0)
    plt.close()

#==============================================================================================
# Validación dentro de 2km
#==============================================================================================
def distancia_haversine(lat1, lon1, lat2, lon2):
    """
    Devuelve la distancia en metros entre dos puntos lat/lon usando Haversine.
    """
    R = 6371000  # radio de la Tierra en metros
    
    lat1, lon1, lat2, lon2 = map(np.radians, [lat1, lon1, lat2, lon2])
    
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    
    a = np.sin(dlat/2)**2 + np.cos(lat1)*np.cos(lat2)*np.sin(dlon/2)**2
    c = 2 * np.arcsin(np.sqrt(a))
    
    return R * c  # distancia en metros


#==============================================================================================
# Configuración del usuario
#==============================================================================================
def obtener_parametros_usuario():
    print("\n=== CONFIGURACIÓN DEL PROYECTO ===")

    codigo = input("👉 Ingrese el CÓDIGO del proyecto: ").strip()
    while codigo == "":
        codigo = input("Código no puede estar vacío. Intente nuevamente: ").strip()

    localidad = input("👉 Ingrese LA LOCALIDAD: ").strip()
    while localidad == "":
        localidad = input("Localidad no puede estar vacía. Intente nuevamente: ").strip()

    tecnologia = input("👉 Ingrese el tipo de tecnología para reconstrucción (UMTS, LTE, NR): ").strip()
    while tecnologia == "":
        tecnologia = input("Tecnología no puede estar vacía. Intente nuevamente: ").strip()

    mnc_raw = input("👉 Ingrese MNC(s) separados por coma (ej: 103,123): ").strip()
    if mnc_raw == "":
        print("⚠ No ingresó MNC, se usará por defecto: 103,123")
        mnc_list = ["103", "123"]
    else:
        mnc_list = [m.strip() for m in mnc_raw.split(",") if m.strip() != ""]

    return codigo, localidad, tecnologia, mnc_list

#==============================================================================================
# Ejecución principal
#==============================================================================================
def main():
    
    print("=== Iniciando procesamiento IMT ===")

    CODIGO, LOCALIDAD, TECNOLOGIA, MNC_SELECCIONADOS = obtener_parametros_usuario()
    inicio = time.time()
    # 1. Leer CSV
    print("Buscando CSV en carpeta datos/ ...")
    input_archivos = sorted(DATOS.glob("*.csv"))
    if not input_archivos:
        print(f"❌ No se encontraron CSV en: {DATOS}")
        return
    else:
        print("Archivos encontrados:")
        for f in input_archivos:
            print("  -", f.name)

    df_list = []
    for f in input_archivos:
        d = leer_y_preparar(f)
        if d is not None:
            df_list.append(d)
            print("✓ Cargado:", f.name)

    df_unidos = pd.concat(df_list, ignore_index=True)

    # Eliminar filas sin coordenadas, quitar espacios, poner como flotantes
    df_unidos = df_unidos.dropna(subset=["Longitude", "Latitude"])
    df_unidos["Latitude"] = pd.to_numeric(df_unidos["Latitude"], errors="coerce")
    df_unidos["Longitude"] = pd.to_numeric(df_unidos["Longitude"], errors="coerce")

    df_unidos["Fila en log"] = df_unidos.index + 1

    carpeta_local = SALIDAS / f"{CODIGO} {LOCALIDAD}"
    ensure_dirs(carpeta_local)

    archivo_logs_unidos = carpeta_local / f"{CODIGO} {LOCALIDAD} Logs Unidos.xlsx"
    safe_save_excel(df_unidos, archivo_logs_unidos)

    # 2. Validación / reconstrucción / enriquecimiento
    df_validados = df_unidos.copy()

    df_validados = reconstruir_globalcellid(df_validados, TECNOLOGIA, tabla_mnc)
    df_validados = renombrar_columnas(df_validados)
    df_validados = establecer_RSRP_RSRQ(df_validados)

    df_validados["Tecnología"] = df_validados["GlobalCellId"].apply(extraer_tecnologia)
    df_validados["MCC-MNC"] = df_validados["GlobalCellId"].apply(extraer_mcc_mnc)
    df_validados["Global CellId /PCI"] = df_validados["GlobalCellId"].apply(
        extraer_globalcell_pci
    )

    df_validados = establecer_prstm(df_validados)
    df_validados = asignar_banda(df_validados)
    df_validados = asignar_punto_medida(df_validados)

    # Reordenar columnas: insertar las nuevas en la posición original de GlobalCellId
    try:
        if "GlobalCellId" in df_validados.columns:
            pos = df_validados.columns.get_loc("GlobalCellId")
            df_validados = df_validados.drop(columns=["GlobalCellId"])

            # Insertar columnas derivadas en el mismo lugar
            if "Tecnología" in df_validados.columns:
                df_validados.insert(pos, "Tecnología", df_validados.pop("Tecnología"))
            if "MCC-MNC" in df_validados.columns:
                df_validados.insert(pos + 1, "MCC-MNC", df_validados.pop("MCC-MNC"))
            if "Global CellId /PCI" in df_validados.columns:
                df_validados.insert(pos + 2, "Global CellId /PCI", df_validados.pop("Global CellId /PCI"))
            if "PRSTM" in df_validados.columns:
                df_validados.insert(pos + 3, "PRSTM", df_validados.pop("PRSTM"))

        if "Punto de medida" in df_validados.columns and "Fila en log" in df_validados.columns:
            pos_punto = df_validados.columns.get_loc("Punto de medida")
            df_validados.insert(pos_punto, "Fila en log", df_validados.pop("Fila en log"))
    except KeyError as e:
        print(f"⚠ Advertencia: No se pudo reorganizar columnas: {e}")
        print("  Continuando con orden actual de columnas.")

    # Reordenar Latitud y Longitud después de Time
    pos_time = df_validados.columns.get_loc("Fecha/hora")
    df_validados.insert(pos_time + 1, "Latitud", df_validados.pop("Latitud"))

    pos_lat = df_validados.columns.get_loc("Latitud")
    df_validados.insert(pos_lat + 1, "Longitud", df_validados.pop("Longitud"))

    df_validados = eliminar_columnas(df_validados)

    archivo_prstm_validados = carpeta_local / f"{CODIGO} {LOCALIDAD} PRSTM_validados.xlsx"
    safe_save_excel(df_validados, archivo_prstm_validados)

    # 3. UT procesar (operador objetivo, mejor muestra por punto)
    df_procesar = df_validados.copy()
    df_procesar = df_procesar[df_procesar["PRSTM"] == OPERADOR_OBJETIVO].copy()
    df_procesar = marcar_mejor_muestra(df_procesar)
    
    # --------------------------------------------------------
    # Crear la columna nueva "dentro_2km"
    # --------------------------------------------------------
    archivo_coord = DATOS / "coordenadas_origen.txt"

    puntos_origen = []

    # Validar que el archivo exista
    if not archivo_coord.exists():
        print(f"❌ Error: No se encontró archivo de coordenadas: {archivo_coord}")
        print("Por favor, cree el archivo 'coordenadas_origen.txt' en la carpeta datos/")
        input("Presione ENTER para terminar...")
        return

    try:
        with open(archivo_coord, "r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea:
                    continue
                
                partes = [x.strip() for x in linea.split(",")]
                
                if len(partes) != 3:
                    print(f"⚠ Línea inválida en archivo de coordenadas: {linea}")
                    continue
                
                nombre, lat, lon = partes
                try:
                    puntos_origen.append({
                        "nombre": nombre,
                        "lat": float(lat),
                        "lon": float(lon)
                    })
                except ValueError as e:
                    print(f"⚠ Error convirtiendo coordenadas en línea '{linea}': {e}")
                    continue
    except Exception as e:
        print(f"❌ Error leyendo archivo de coordenadas: {e}")
        input("Presione ENTER para terminar...")
        return
    
    # Validar que se hayan leído al menos 2 puntos
    if len(puntos_origen) < 2:
        print(f"❌ Error: Se requieren al menos 2 puntos en el archivo de coordenadas.")
        print(f"   Solo se encontraron: {len(puntos_origen)} punto(s)")
        input("Presione ENTER para terminar...")
        return
    
    nombre0 = puntos_origen[0]["nombre"]
    lat0 = puntos_origen[0]["lat"]
    lon0 = puntos_origen[0]["lon"]
    
    nombre1 = puntos_origen[1]["nombre"]
    lat1 = puntos_origen[1]["lat"]
    lon1 = puntos_origen[1]["lon"]
    
    df_procesar["dist_m"] = distancia_haversine(df_procesar["Latitud"], df_procesar["Longitud"], lat0, lon0)

    df_procesar["dentro_2km"] = np.where(df_procesar["dist_m"] <= 2000, " ", "Por fuera del área")

    resumen = generar_resumen_rsrp(df_procesar)
    
    # 4. KMZ y mapa (solo muestras conservadas)
    df_conservada = df_procesar[
        (df_procesar["Exclusiones"] == " ") &
        (df_procesar["dentro_2km"] == " ")
    ].copy()
    
    
    generar_kmz_completo(df_conservada, CODIGO, LOCALIDAD, MNC_SELECCIONADOS)
    radio_m = 2000
    #radio_m = generar_radio_2km(CODIGO, LOCALIDAD)
    #generar_kmz_potencia(df_conservada, CODIGO, LOCALIDAD)
    #generar_kmz_calidad(df_conservada, CODIGO, LOCALIDAD)
    #generar_kmz_banda(df_conservada, CODIGO, LOCALIDAD)
    #generar_kmz_mnc(df_conservada, CODIGO, LOCALIDAD, MNC_SELECCIONADOS)
    #generar_kmz_globalcellpci(df_conservada, CODIGO, LOCALIDAD)
    
    generar_mapa_rsrp(df_conservada, CODIGO, LOCALIDAD, 
    lat0, lon0, nombre0, 
    lat1, lon1, nombre1, 
    radio_m)
    generar_mapa_rsrq(df_conservada, CODIGO, LOCALIDAD,
    lat0, lon0, nombre0, 
    lat1, lon1, nombre1, 
    radio_m)
    generar_mapa_banda(df_conservada, CODIGO, LOCALIDAD,
    lat0, lon0, nombre0, 
    lat1, lon1, nombre1, 
    radio_m)
    generar_mapa_mnc(df_conservada, CODIGO, LOCALIDAD, MNC_SELECCIONADOS,
    lat0, lon0, nombre0, 
    lat1, lon1, nombre1, 
    radio_m)
    generar_mapa_globalcellpci(df_conservada, CODIGO, LOCALIDAD,
    lat0, lon0, nombre0, 
    lat1, lon1, nombre1, 
    radio_m)

    # 5. Exportar Excel final con resumen

    df_procesar.rename(columns={
        "Exclusiones": "Mejor Medida",
        "dentro_2km": "Exclusiones",
        "Global CellId /PCI": "CID sector"
    }, inplace=True)
  
    pos_banda = df_procesar.columns.get_loc("Banda (MHz)")
    df_procesar.insert(pos_banda + 1, "Exclusiones", df_procesar.pop("Exclusiones"))
    
    archivo_salida = carpeta_local / f"Ampliación Cobertura {CODIGO} {LOCALIDAD}.xlsx"
    exportar_excel_con_resumen(df_procesar, resumen, archivo_salida)

    print("✔ Archivo Excel con datos UT + resumen generado.")

    fin = time.time()
    duracion = fin - inicio

    print("\n======================================")
    print(f"Tiempo total de ejecución: {duracion:.2f} segundos")
    print("======================================\n")

    input("Presione ENTER para terminar...")

#==============================================================================================
if __name__ == "__main__":
    main()
